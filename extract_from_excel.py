#!/usr/bin/env python3
"""Extract accounting data from Honeybird Excel file and generate dashboard JSON files."""

import json
import os
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

EXCEL_PATH = "/sessions/vigilant-bold-faraday/mnt/1. Monthly Accounting/1. Entities/Honeybird/Accounting File/Honeybird_Accounting_FY2026.xlsx"
DATA_DIR = "/sessions/vigilant-bold-faraday/honeybird-dashboard/data"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def parse_date(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s


# ========== TRANSACTIONS ==========
def extract_transactions():
    transactions = []
    for name in wb.sheetnames:
        if not name.startswith('Txn'):
            continue
        ws = wb[name]
        name_lower = name.lower()

        # Determine bank and account from sheet name
        if 'bank zer' in name_lower:
            bank = 'Bank Zero'
        elif 'capitec' in name_lower:
            bank = 'Capitec'
        elif 'easyequi' in name_lower:
            bank = 'EasyEquities'
        else:
            bank = 'Unknown'

        # Account number from row 4
        acct_num = ''
        for row in ws.iter_rows(min_row=4, max_row=4, values_only=True):
            text = str(row[0] or '')
            m = re.search(r'(\d{4,})', text)
            if m:
                acct_num = m.group(1)

        # Row 6 = headers: #, Post Date, Trans Date, Description, Reference, Debit (Out), Credit (In), Balance, Proposed Category, BS / IS Allocation
        for row in ws.iter_rows(min_row=7, values_only=True):
            cells = list(row)
            # Skip empty/header/total rows
            if not cells[0] or not isinstance(cells[0], (int, float)):
                continue
            post_date = cells[1]
            if not post_date:
                continue

            date_str = parse_date(post_date)
            description = str(cells[3] or '')
            reference = str(cells[4] or '')
            debit = float(cells[5]) if cells[5] else 0
            credit = float(cells[6]) if cells[6] else 0
            balance_val = float(cells[7]) if cells[7] else 0
            category = str(cells[8] or '')
            allocation = str(cells[9] or '') if len(cells) > 9 else ''

            amount = credit - debit
            if amount == 0:
                continue

            txn = {
                "date": date_str,
                "bank": bank,
                "account": acct_num,
                "entity": "Honeybird",
                "description": description,
                "reference": reference,
                "amount": round(amount, 2),
                "balance": round(balance_val, 2),
                "category": category if category else ('Other Income' if amount > 0 else 'Other Expenses'),
                "allocation": allocation
            }
            transactions.append(txn)

    transactions.sort(key=lambda t: t.get('date', ''))
    return transactions


# ========== ENTITY REGISTRY ==========
def extract_entity_info():
    ws = wb['Entity Info']
    data = {}
    bank_accounts = []

    for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
        cells = [str(c).strip() if c else '' for c in row]
        if cells[0] and cells[1]:
            data[cells[0]] = cells[1]
        # Bank account rows
        if any(kw in cells[0].lower() for kw in ['bank zero', 'capitec', 'easyequities']):
            bank_accounts.append({
                "bank": cells[0],
                "account_number": cells[1],
                "account_type": cells[2] if cells[2] else "Business Account",
                "branch": cells[3] if len(cells) > 3 and cells[3] else ""
            })

    vat_status = data.get('VAT Registration', data.get('VAT Status', data.get('VAT', '')))
    vat_registered = vat_status.lower() not in ('not registered', 'no', 'n/a', '')

    if not bank_accounts:
        bank_accounts = [
            {"bank": "Bank Zero", "account_number": "80204693662", "account_type": "Business Account", "branch": ""},
            {"bank": "Capitec", "account_number": "1053469489", "account_type": "Business Account", "branch": "450105"},
            {"bank": "Capitec", "account_number": "1054023212", "account_type": "Business Account", "branch": "450105"},
            {"bank": "EasyEquities", "account_number": "3212", "account_type": "Investment Account", "branch": ""}
        ]

    return {
        "Honeybird": {
            "legal_name": data.get("Entity Name", data.get("Company Name", "HONEYBIRD PRIVATE EQUITY (PTY) LTD")),
            "entity_type": data.get("Entity Type", "Private Company"),
            "registration_number": data.get("Registration Number", "2024/447842/07"),
            "vat_number": data.get("VAT Number", ""),
            "vat_registered": vat_registered,
            "income_tax_number": data.get("Income Tax Number", ""),
            "address": data.get("Address", data.get("Registered Address", "151 Meyer Street, Loeriepark, Western Cape, 6529")),
            "directors": [data.get("Directors", "")] if data.get("Directors") else [],
            "shareholders": [],
            "financial_year_end": "April",
            "date_of_registration": data.get("Date of Registration", "22/07/2024"),
            "previous_names": ["Entwood (Pty) Ltd", "Entwood Investment Group"],
            "bank_accounts": bank_accounts
        }
    }


# ========== BANK RECONCILIATION ==========
def extract_recon():
    ws = wb['Bank Reconstruction']
    accounts = {}
    current_acct_num = None
    current_bank = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = [str(c).strip() if c else '' for c in row]
        line = ' '.join(cells).lower()

        # Detect account headers
        if 'bank zero' in line:
            current_bank = 'Bank Zero'
            m = re.search(r'(\d{8,})', line)
            current_acct_num = m.group(1) if m else '80204693662'
        elif 'capitec' in line and re.search(r'\d{8,}', line):
            current_bank = 'Capitec'
            m = re.search(r'(\d{8,})', line)
            current_acct_num = m.group(1) if m else ''
        elif 'easyequities' in line or 'easy equities' in line:
            current_bank = 'EasyEquities'
            current_acct_num = '3212'

        if current_acct_num and current_acct_num not in accounts:
            accounts[current_acct_num] = {
                "bank": current_bank,
                "account_type": "Investment Account" if current_bank == 'EasyEquities' else "Business Account",
                "statement_closing_balance": 0,
                "calculated_balance": 0,
                "opening_balance": 0,
                "reconciled": True,
                "difference": 0,
                "transaction_count": 0,
                "latest_statement_date": ""
            }

        if current_acct_num:
            for i, cell_str in enumerate(cells):
                cl = cell_str.lower()
                if 'closing' in cl and 'balance' in cl:
                    for j in range(i + 1, len(row)):
                        if row[j] is not None and isinstance(row[j], (int, float)):
                            accounts[current_acct_num]["statement_closing_balance"] = round(float(row[j]), 2)
                            accounts[current_acct_num]["calculated_balance"] = round(float(row[j]), 2)
                            break
                elif 'opening' in cl and 'balance' in cl:
                    for j in range(i + 1, len(row)):
                        if row[j] is not None and isinstance(row[j], (int, float)):
                            accounts[current_acct_num]["opening_balance"] = round(float(row[j]), 2)
                            break
                elif 'difference' in cl:
                    for j in range(i + 1, len(row)):
                        if row[j] is not None and isinstance(row[j], (int, float)):
                            val = round(float(row[j]), 2)
                            accounts[current_acct_num]["difference"] = val
                            accounts[current_acct_num]["reconciled"] = abs(val) < 0.01
                            break

    return {"accounts": {"Honeybird": accounts}}


# ========== TRIAL BALANCE ==========
def extract_trial_balance():
    ws = wb['Trial Balance']
    tb = []
    section = ''
    # Data starts at row 7 (row 6 is header)
    for row in ws.iter_rows(min_row=7, values_only=True):
        cells = list(row)
        # Column layout: [None, Account Code, Description, Debit, Credit, WP Ref]
        acct_code = cells[1] if len(cells) > 1 else None
        description = str(cells[2] or '') if len(cells) > 2 else ''
        debit = float(cells[3]) if len(cells) > 3 and isinstance(cells[3], (int, float)) else 0
        credit = float(cells[4]) if len(cells) > 4 and isinstance(cells[4], (int, float)) else 0
        wp_ref = str(cells[5] or '') if len(cells) > 5 else ''

        # Section headers (no account code, but have description)
        if not acct_code and description in ('ASSETS', 'EQUITY', 'LIABILITIES', 'INCOME', 'EXPENSES'):
            section = description
            continue

        if acct_code:
            tb.append({
                "account_code": str(acct_code),
                "account_name": description,
                "section": section,
                "debit": round(debit, 2),
                "credit": round(credit, 2),
                "wp_ref": wp_ref
            })
    return tb


# ========== FINANCIAL STATEMENTS ==========
def extract_financial_statements():
    ws = wb['Financial Statements']
    fs = {
        "balance_sheet": {
            "assets": [],
            "total_assets": 0,
            "equity": [],
            "total_equity": 0,
            "liabilities": [],
            "total_liabilities": 0,
            "total_equity_liabilities": 0,
            "out_of_balance": 0
        },
        "income_statement": {
            "expenses": [],
            "net_loss": 0
        }
    }

    section = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = list(row)
        desc = str(cells[1] or '').strip() if len(cells) > 1 else ''
        amount = cells[2] if len(cells) > 2 and isinstance(cells[2], (int, float)) else None
        desc_lower = desc.lower()

        if 'statement of financial position' in desc_lower:
            section = 'bs'
        elif 'statement of comprehensive income' in desc_lower:
            section = 'is'

        if section == 'bs':
            if desc_lower == 'assets' or desc_lower == 'current assets':
                continue
            elif 'equity' in desc_lower and 'total' not in desc_lower and 'liabilities' not in desc_lower:
                continue
            elif desc_lower.startswith('non-current') or desc_lower.startswith('current'):
                continue

            if desc.startswith('  ') and amount is not None:
                name = desc.strip()
                if amount is not None:
                    # Determine section by position
                    if 'cash' in name.lower() or 'bank' in name.lower():
                        fs['balance_sheet']['assets'].append({"name": name, "amount": round(float(amount), 2)})
                    elif 'share' in name.lower() or 'contribution' in name.lower() or 'accumulated' in name.lower() or 'retained' in name.lower():
                        fs['balance_sheet']['equity'].append({"name": name, "amount": round(float(amount), 2)})
                    elif 'loan' in name.lower() or 'interaccount' in name.lower():
                        fs['balance_sheet']['liabilities'].append({"name": name, "amount": round(float(amount), 2)})

            if 'total assets' in desc_lower and amount is not None:
                fs['balance_sheet']['total_assets'] = round(float(amount), 2)
            elif 'total equity' == desc_lower.strip() and amount is not None:
                fs['balance_sheet']['total_equity'] = round(float(amount), 2)
            elif 'total liabilities' in desc_lower and amount is not None:
                fs['balance_sheet']['total_liabilities'] = round(float(amount), 2)
            elif 'total equity and liabilities' in desc_lower and amount is not None:
                fs['balance_sheet']['total_equity_liabilities'] = round(float(amount), 2)
            elif 'out of balance' in desc_lower and amount is not None:
                fs['balance_sheet']['out_of_balance'] = round(float(amount), 2)

        elif section == 'is':
            if desc.startswith('  ') or (desc and not desc.startswith('EXPENSES') and not desc.startswith('STATEMENT')):
                name = desc.strip()
                if amount is not None and name:
                    if 'loss for' in name.lower() or 'profit for' in name.lower():
                        fs['income_statement']['net_loss'] = round(float(amount), 2)
                    else:
                        fs['income_statement']['expenses'].append({"name": name, "amount": round(float(amount), 2)})

    return fs


# ========== WORKING PAPERS ==========
def extract_working_papers():
    ws = wb['Working Papers']
    schedules = []
    current = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = list(row)
        label = str(cells[0] or '').strip() if cells[0] else ''
        desc = str(cells[1] or '').strip() if len(cells) > 1 else ''
        amount = cells[2] if len(cells) > 2 and isinstance(cells[2], (int, float)) else None

        if label.startswith('WP -'):
            if current:
                schedules.append(current)
            current = {"name": label, "entries": []}
        elif current and desc and amount is not None:
            current["entries"].append({"description": desc, "amount": round(float(amount), 2)})

    if current:
        schedules.append(current)
    return schedules


# ========== INVOICES ==========
def extract_invoices():
    ws = wb['Invoices']
    invoices = []
    # Row 6 = headers: #, Invoice Date, Invoice Number, Supplier/Vendor, Description, Amount, VAT, Excl VAT, Category, Matched, Bank Txn Date, Source File, BS / IS
    for row in ws.iter_rows(min_row=7, values_only=True):
        cells = list(row)
        if not cells[0] or not isinstance(cells[0], (int, float)):
            continue

        inv = {
            "entity": "Honeybird",
            "invoice_date": parse_date(cells[1]) if cells[1] else "",
            "invoice_number": str(cells[2] or ''),
            "vendor": str(cells[3] or ''),
            "description": str(cells[4] or ''),
            "amount_incl_vat": float(cells[5]) if cells[5] and isinstance(cells[5], (int, float)) else 0,
            "vat_amount": float(cells[6]) if cells[6] and isinstance(cells[6], (int, float)) else 0,
            "amount_excl_vat": float(cells[7]) if cells[7] and isinstance(cells[7], (int, float)) else 0,
            "category": str(cells[8] or ''),
            "matched": str(cells[9] or '').lower() in ('yes', 'true', 'matched'),
            "bank_txn_date": parse_date(cells[10]) if cells[10] else "",
            "filename": str(cells[11] or ''),
            "allocation": str(cells[12] or '') if len(cells) > 12 else '',
            "file_type": ".pdf",
            "status": "Filed"
        }
        invoices.append(inv)
    return invoices


# ========== QUERIES ==========
def extract_queries():
    ws = wb['Query List']
    queries = []
    header_found = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = [str(c).strip() if c else '' for c in row]
        if not header_found:
            if any('#' in c or 'query' in c.lower() for c in cells):
                header_found = True
            continue
        if cells[0] and cells[0] != '':
            queries.append({
                "number": cells[0],
                "description": cells[1] if len(cells) > 1 else '',
                "status": cells[2] if len(cells) > 2 else 'Open',
                "assigned_to": cells[3] if len(cells) > 3 else '',
                "notes": cells[4] if len(cells) > 4 else ''
            })
    return queries


# ========== RUN ==========
print("=== Extracting all data ===\n")

registry = extract_entity_info()
transactions = extract_transactions()
recon = extract_recon()
trial_balance = extract_trial_balance()
fin_stmts = extract_financial_statements()
working_papers = extract_working_papers()
invoices = extract_invoices()
queries = extract_queries()

# DERIVE BALANCE SHEET LINE ITEMS FROM TRIAL BALANCE (more reliable than parsing FS sheet)
bs = fin_stmts['balance_sheet']
if not bs['assets'] or not bs['equity'] or not bs['liabilities']:
    print("  -> Deriving BS line items from Trial Balance...")
    bs['assets'] = []
    bs['equity'] = []
    bs['liabilities'] = []
    total_assets = 0
    total_equity = 0
    total_liabilities = 0
    for tb in trial_balance:
        net = round(tb['debit'] - tb['credit'], 2)
        if tb['section'] == 'ASSETS':
            bs['assets'].append({"name": tb['account_name'], "amount": round(tb['debit'] - tb['credit'], 2)})
            total_assets += net
        elif tb['section'] == 'EQUITY':
            # Equity: credit balance positive, debit balance (losses) negative
            eq_amt = round(tb['credit'] - tb['debit'], 2)
            bs['equity'].append({"name": tb['account_name'], "amount": eq_amt})
            total_equity += eq_amt
        elif tb['section'] == 'LIABILITIES':
            # Liabilities: credit balance positive, debit balance negative (receivable)
            li_amt = round(tb['credit'] - tb['debit'], 2)
            bs['liabilities'].append({"name": tb['account_name'], "amount": li_amt})
            total_liabilities += li_amt
    bs['total_assets'] = round(total_assets, 2)
    bs['total_equity'] = round(total_equity, 2)
    bs['total_liabilities'] = round(total_liabilities, 2)
    bs['total_equity_liabilities'] = round(total_equity + total_liabilities, 2)
    bs['out_of_balance'] = round(total_assets - (total_equity + total_liabilities), 2)

# Also ensure IS has revenue section
if 'revenue' not in fin_stmts['income_statement']:
    fin_stmts['income_statement']['revenue'] = []
    fin_stmts['income_statement']['total_revenue'] = 0
    fin_stmts['income_statement']['total_expenses'] = fin_stmts['income_statement']['net_loss']

# Update recon with transaction counts
for t in transactions:
    acct = t.get('account', '')
    if acct in recon['accounts'].get('Honeybird', {}):
        recon['accounts']['Honeybird'][acct]['transaction_count'] += 1
        if t['date'] > recon['accounts']['Honeybird'][acct].get('latest_statement_date', ''):
            recon['accounts']['Honeybird'][acct]['latest_statement_date'] = t['date']

# Print summary
print(f"Transactions: {len(transactions)}")
for t in transactions:
    print(f"  {t['date']} | {t['bank']} {t['account']} | {t['description'][:35]:35} | {t['amount']:>12,.2f} | {t['category']}")

print(f"\nRecon accounts: {len(recon['accounts']['Honeybird'])}")
for acct, info in recon['accounts']['Honeybird'].items():
    print(f"  {info['bank']} {acct}: closing={info['statement_closing_balance']:,.2f} reconciled={info['reconciled']} txns={info['transaction_count']}")

print(f"\nTrial Balance: {len(trial_balance)} entries")
for tb in trial_balance:
    print(f"  {tb['account_code']:>5} | {tb['section']:12} | {tb['account_name'][:35]:35} | Dr: {tb['debit']:>12,.2f} | Cr: {tb['credit']:>12,.2f}")

print(f"\nFinancial Statements:")
print(f"  BS Assets: {fin_stmts['balance_sheet']['total_assets']:,.2f}")
print(f"  BS Equity: {fin_stmts['balance_sheet']['total_equity']:,.2f}")
print(f"  BS Liabilities: {fin_stmts['balance_sheet']['total_liabilities']:,.2f}")
print(f"  Out of balance: {fin_stmts['balance_sheet']['out_of_balance']:,.2f}")
print(f"  IS Net loss: {fin_stmts['income_statement']['net_loss']:,.2f}")

print(f"\nWorking Papers: {len(working_papers)} schedules")
for wp in working_papers:
    print(f"  {wp['name']}: {len(wp['entries'])} entries")

print(f"\nInvoices: {len(invoices)}")
print(f"Queries: {len(queries)}")

# ========== SAVE ==========
print("\n=== Saving JSON files ===")
files = {
    'entity_registry.json': registry,
    'transactions.json': transactions,
    'recon.json': recon,
    'invoices.json': invoices,
    'trial_balance.json': trial_balance,
    'financial_statements.json': fin_stmts,
    'working_papers.json': working_papers,
    'queries.json': queries
}

for fname, data in files.items():
    path = os.path.join(DATA_DIR, fname)
    with open(path, 'w') as f:
        json.dump(data, f, indent=2, default=str)
    print(f"  Saved {fname}")

print("\n=== DONE ===")
